import openpyxl as xcl	#importing module


wb = xcl.load_workbook('expenses.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
	cell = sheet.cell(row, 3)
	corrected_price = cell.value * 0.9
	corrected_price_cell = sheet.cell(row, 4)
	corrected_price_cell.value = corrected_price

wb.save('expenses2.xlsx')